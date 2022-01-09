import pandas as pd
import csv


from openpyxl import Workbook, load_workbook

wb = load_workbook('employeedata.xlsx')

ws = wb.active


range = ws["B2":"B30"]

def employee():
  for cell in range:
    for x in cell:
      print(x.value)
employee()


def database():
  for cell in range:
    for x in cell:
      text = x.value
      changeSufix = text.replace("helpinghands.cm","handsinhands.org")
      x.value = changeSufix
      print(changeSufix)
      wb.save('updatedfile.xlsx')
      

database()


df = pd.read_csv("employeedata.csv")

print(df)


df['Emails'] = df['Emails'].str.replace('helpinghands.cm', 'handsinhands.org')
print(df)

df.to_csv("updated.csv", index=False)